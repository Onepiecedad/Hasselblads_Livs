#!/usr/bin/env python3
"""Extract Axels lista to JSON for migration"""
import openpyxl
import json
import os

wb = openpyxl.load_workbook(os.path.join(os.path.dirname(__file__), '..', 'Axels lista (3) (1).xlsx'))
sheet = wb.active

products = []
for row_idx in range(2, sheet.max_row + 1):
    name = sheet.cell(row=row_idx, column=1).value
    if not name:
        continue
    
    vikt = sheet.cell(row=row_idx, column=2).value
    kgst = sheet.cell(row=row_idx, column=3).value
    active = sheet.cell(row=row_idx, column=4).value
    price = sheet.cell(row=row_idx, column=5).value
    sale_price = sheet.cell(row=row_idx, column=6).value
    
    products.append({
        'name': str(name).strip(),
        'vikt_gram': str(vikt).strip() if vikt else '',
        'kg_st': str(kgst).strip() if kgst else 'st',
        'active': int(active) if active is not None else 0,
        'price': float(price) if price else None,
        'sale_price': float(sale_price) if sale_price else None,
    })

# Write JSON
out_path = os.path.join(os.path.dirname(__file__), 'axels-data.json')
with open(out_path, 'w', encoding='utf-8') as f:
    json.dump(products, f, ensure_ascii=False, indent=2)

print(f"Exported {len(products)} products to {out_path}")

# Summary
kg_products = [p for p in products if p['kg_st'].lower() == 'kg']
st_with_vikt = [p for p in products if p['kg_st'].lower() == 'st' and p['vikt_gram']]
st_without_vikt = [p for p in products if p['kg_st'].lower() == 'st' and not p['vikt_gram']]

print(f"\nKg-products: {len(kg_products)}")
for p in kg_products:
    print(f"  {p['name']}: {p['price']} kr/kg, vikt: {p['vikt_gram']}")
print(f"\nSt with weight: {len(st_with_vikt)}")
print(f"St without weight: {len(st_without_vikt)}")
